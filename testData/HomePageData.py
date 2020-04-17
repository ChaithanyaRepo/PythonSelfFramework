import openpyxl

from utilities.BaseClass import BaseClass


class HomePageData(BaseClass):

    test_HomePage_Data = [
        {"Name": "Chaithanya", "Email": "nayakchaithanya12@yahoo.com", "Password": "Chaisush@2010", "Gender": "Male",
         "DOB": "13/12/1994"},
        {"Name": "Sushma", "Email": "sush.sprabhu@gmail.com", "Password": "Chaisush@2010", "Gender": "Female",
         "DOB": "20/10/1997"},
        {"Name": "Anusha", "Email": "sush.sprabhu@gmail.com", "Password": "Chaisush@2010", "Gender": "Female",
         "DOB": "24/02/1993"}]

    # To avoid the creation of class object
    @staticmethod
    def getTestData(TestCaseName):

        HomePageData_Dict = {}
        book = openpyxl.load_workbook("/home/chaitanya/Documents/PythonDemo.xlsx")
        for i in range(1, book.active.max_row + 1):  # to get rows
            if book.active.cell(row=i, column=1).value == TestCaseName:
                for j in range(2, book.active.max_column + 1):  # to get columns
                    HomePageData_Dict[book.active.cell(row=1, column=j).value] = book.active.cell(row=i, column=j).value
            elif TestCaseName == "":
                for j in range(2, book.active.max_column + 1):  # to get columns
                    HomePageData_Dict[book.active.cell(row=1, column=j).value] = book.active.cell(row=i, column=j).value

        return [HomePageData_Dict]
